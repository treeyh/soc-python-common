#-*- encoding: utf-8 -*-

# https://jira.readthedocs.io/en/latest/examples.html#issues
# pip install jira

from jira import JIRA
from openpyxl import Workbook, load_workbook


jira_url = ''
jira_username = ''
jira_password = ''
jira = None


excel_path = ''


# id   issue id
# project      所属项目
# issuetype   类型
# summary 概述
# description  描述
# priority 优先级
# labels   标签
# assignee  经办人
# customfield_10300 sprint
# customfield_10804 到期日
# customfield_10308 故事点
_column_list = ['id', 'project', 'issuetype', 'summary', 'description', 'priority', 'labels', 'assignee', 'customfield_10300', 'customfield_10804', 'customfield_10308']

_wb = None
_ws = None

def load_issues():
  global excel_path, _ws
  _wb = load_workbook(filename = excel_path)
  _ws = _wb.active
  rows = _ws.values
  issues = []

  rowIndex = 0
  for row in rows:
    rowIndex += 1
    id = row[0]
    project = row[1]
    issuetype = row[2]
    summary = row[3]
    description = row[4]
    priority = row[5]
    labels = row[6]
    assignee = row[7]
    customfield_10300 = row[8]
    customfield_10804 = row[9]
    customfield_10308 = row[10]

    if project == None or issuetype == None or summary == None or priority == None or assignee == None or rowIndex == 1:
      continue
    issue = {
      'row': rowIndex,
      'issue': {
        'project': {'key': project},
        'issuetype': {'name': issuetype},
        'summary': summary,
        'description' : description,
        'priority' : {'name': priority},
        'labels' : labels.split(',') if labels != None else [],
        'assignee': {'name': assignee},
      }
    }
    if id != None and id.strip() != '':
      issue['issue']['id'] = id
    if customfield_10300 != None and customfield_10300.strip() != '':
      issue['issue']['customfield_10300'] = {'name': customfield_10300}
    if customfield_10804 != None:
      issue['issue']['customfield_10804'] = customfield_10804.strftime('%Y-%m-%d') if customfield_10804 != None else None
    if customfield_10308 != None:
      issue['issue']['customfield_10308'] = customfield_10308

    issues.append(issue)
  return issues



def create_or_update_issues(issues):
  global jira, _ws,_wb,excel_path

  last_parent_key = ''
  for issue in issues:
    if issue['issue'].get('id', None) == None:
      # create
      if issue['issue']['issuetype']['name'] == '子任务':
        if last_parent_key != '':
          issue['issue']['parent'] = {'key': last_parent_key}
        else:
          continue
      new = jira.create_issue(fields=issue['issue'])
      if issue['issue']['issuetype']['name'] == '故事':
        last_parent_key = new.key
      # 回写excel issue id
      c = _ws.cell(row = issue['row'], column = 1)
      c.value = new.key
      print('create issue over:'+new.key)
    else:
      issue = jira.issue(issue['issue']['id'])
      del(issue['issue']['id'])
      issue.update(fields=issue['issue'])
      print('update issue over:'+new.key)
  _wb.save(excel_path)
      
      


def run():

  global jira_url, jira_username, jira_password, jira
  issues = load_issues()

  if len(issues) <= 0:
    print('not issues')
    return

  jira = JIRA(server=jira_url, basic_auth=(jira_username, jira_password))
  
  create_or_update_issues(issues)
  

  # RTAN
  # project = jira.project('RTAN')
  
  # print(project)
  # print(project.id)
  # print(project.name)
  # print('----------------')
  
  # issue = jira.issue('RTAN-114')
  # for key in dir(issue.fields):
  #   print(key + ':')
  #   eval('print(issue.fields.'+key+')')

  


if __name__ == '__main__':
  run()    



